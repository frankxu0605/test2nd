"""Excel file import endpoint for all modules."""

from datetime import date, datetime, time as time_type
from decimal import Decimal
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from sqlalchemy import func

from ..database import get_db
from ..models import Customer, Order, RepaymentPlan, WarehouseEntry, Appointment, Expense, User
from ..schemas import (
    CustomerCreate, OrderCreate, RepaymentPlanCreate,
    WarehouseEntryCreate, AppointmentCreate, ExpenseCreate,
)
from ..auth_deps import get_active_user
from .orders import auto_generate_plans

router = APIRouter(prefix="/api/import", tags=["import"])

# ---- Column mappings: Chinese header -> field name ----
COLUMN_MAPPINGS = {
    "customers": {
        "客户编号": "customer_no",
        "姓名": "name",
        "电话": "phone",
        "身份证": "id_card",
        "地址": "address",
        "邮箱": "email",
        "客户经理": "account_manager",
        "紧急联系人": "emergency_contact",
        "当前逾期": "has_overdue",
        "有房产": "has_property",
    },
    "orders": {
        "日期": "order_date",
        "订单编号": "order_no",
        "客户姓名": "customer_name_import",
        "电话": "phone",
        "身份证": "id_card",
        "地址": "address",
        "邮箱": "email",
        "客户经理": "account_manager",
        "操作员": "operator",
        "紧急联系人": "emergency_contact",
        "克重": "weight",
        "单价": "unit_price",
        "加工费": "processing_fee",
        "公证费": "notary_fee",
        "首付比例": "down_payment_ratio",
        "首付金额": "down_payment",
        "收款账户": "payment_account",
        "分期期数": "installment_periods",
        "每期金额": "installment_amount",
        "状态": "status",
    },
    "repayments": {
        "订单ID": "order_id",
        "期数": "period_no",
        "到期日": "due_date",
        "本金": "principal",
        "利息": "interest",
        "应还金额": "total_amount",
        "已还金额": "paid_amount",
        "还款日期": "paid_date",
        "收款账户": "payment_account",
        "状态": "status",
    },
    "warehouse": {
        "编号": "item_no",
        "条码": "barcode",
        "克重": "weight",
        "单价": "unit_price",
        "总价": "total_price",
        "入库时间": "entry_date",
        "入库员": "entry_operator",
        "出库时间": "exit_date",
        "出库员": "exit_operator",
        "买家": "buyer",
        "销售": "salesperson",
        "备注": "notes",
    },
    "appointments": {
        "客户ID": "customer_id",
        "电话": "phone",
        "预约日期": "appointment_date",
        "预约时间": "appointment_time",
        "事由": "purpose",
        "状态": "status",
        "备注": "notes",
    },
    "expenses": {
        "日期": "expense_date",
        "采购单号": "purchase_order_no",
        "供应商名称": "supplier_name",
        "供应商电话": "supplier_phone",
        "供应商地址": "supplier_address",
        "产品名称": "product_name",
        "支出类别": "category",
        "单位": "unit",
        "数量": "quantity",
        "单价": "unit_price",
        "总价": "total_price",
        "收货人": "receiver",
        "收货电话": "receiver_phone",
        "收货地址": "receiver_address",
        "备注": "notes",
        "支出账户": "payment_account",
    },
}

MODEL_MAP = {
    "customers": Customer,
    "orders": Order,
    "repayments": RepaymentPlan,
    "warehouse": WarehouseEntry,
    "appointments": Appointment,
    "expenses": Expense,
}

SCHEMA_MAP = {
    "customers": CustomerCreate,
    "orders": OrderCreate,
    "repayments": RepaymentPlanCreate,
    "warehouse": WarehouseEntryCreate,
    "appointments": AppointmentCreate,
    "expenses": ExpenseCreate,
}

REQUIRED_FIELDS = {
    "customers": ["name"],
    "orders": ["customer_name_import"],
    "repayments": ["order_id", "period_no", "due_date", "principal", "interest", "total_amount"],
    "warehouse": [],
    "appointments": ["customer_id", "appointment_date", "appointment_time"],
    "expenses": ["expense_date"],
}

MODULE_TITLES = {
    "customers": "客户",
    "orders": "订单",
    "repayments": "还款计划",
    "warehouse": "入库记录",
    "appointments": "预约",
    "expenses": "支出",
}

# Integer fields that Excel may read as float (e.g. 1.0 -> 1)
INT_FIELDS = {"customer_id", "order_id", "period_no", "installment_periods", "quantity", "customer_no"}


def _preprocess_row(module: str, row_data: dict, db: Session, row_idx: int) -> dict:
    """Convert Excel cell values to types expected by Pydantic schemas."""
    for key in list(row_data.keys()):
        val = row_data[key]

        # datetime -> date for date fields
        if isinstance(val, datetime):
            if key == "appointment_time":
                row_data[key] = val.time()
            else:
                row_data[key] = val.date()
        elif isinstance(val, time_type):
            pass  # already correct type

        # Excel reads integers as floats
        if key in INT_FIELDS and isinstance(val, float):
            row_data[key] = int(val)

        # Convert payment_account to string
        if key == "payment_account" and isinstance(val, (int, float)):
            row_data[key] = str(int(val)) if val == int(val) else str(val)

        # Empty string -> None for optional date/time fields
        if key in ("entry_date", "exit_date", "paid_date", "order_date",
                    "appointment_date", "expense_date", "appointment_time") and val == "":
            row_data[key] = None

        # Convert phone/id_card numbers to string
        if key in ("phone", "id_card", "barcode", "item_no", "order_no",
                    "purchase_order_no", "supplier_phone", "receiver_phone"):
            if isinstance(val, (int, float)):
                row_data[key] = str(int(val)) if val == int(val) else str(val)

    # Auto-generate order_no for orders
    if module == "orders":
        if not row_data.get("order_no"):
            row_data["order_no"] = "ORD" + datetime.now().strftime("%Y%m%d%H%M%S%f") + str(row_idx)
        if not row_data.get("order_date"):
            row_data["order_date"] = date.today()

    # Auto-generate purchase_order_no for expenses
    if module == "expenses":
        if not row_data.get("purchase_order_no"):
            today_str = date.today().strftime("%Y%m%d")
            prefix = f"CG{today_str}"
            last = (
                db.query(Expense)
                .filter(Expense.purchase_order_no.like(f"{prefix}%"))
                .order_by(Expense.purchase_order_no.desc())
                .first()
            )
            seq = (int(last.purchase_order_no[-3:]) + 1) if (last and last.purchase_order_no) else 1
            row_data["purchase_order_no"] = f"{prefix}{seq:03d}"

    return row_data


@router.post("/{module}")
async def import_excel(module: str, file: UploadFile = File(...), user: User = Depends(get_active_user), db: Session = Depends(get_db)):
    if module not in COLUMN_MAPPINGS:
        raise HTTPException(status_code=400, detail=f"不支持的模块: {module}")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传Excel文件(.xlsx)")

    contents = await file.read()
    try:
        wb = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="无法读取Excel文件，请确认格式正确")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise HTTPException(status_code=400, detail="Excel文件为空")

    # Build column index -> field name mapping from header row
    headers = rows[0]
    col_map = COLUMN_MAPPINGS[module]
    reverse_map = {}
    for idx, header in enumerate(headers):
        if header and str(header).strip() in col_map:
            reverse_map[idx] = col_map[str(header).strip()]

    if not reverse_map:
        raise HTTPException(status_code=400, detail="未识别到有效的列标题，请使用正确的模板")

    ModelClass = MODEL_MAP[module]
    SchemaClass = SCHEMA_MAP[module]
    required = REQUIRED_FIELDS[module]
    results = {"success": 0, "total": 0, "errors": []}

    # Pre-calculate next customer_no for batch import
    if module == "customers":
        max_no = db.query(func.max(Customer.customer_no)).filter(
            Customer.tenant_id == user.tenant_id
        ).scalar()
        _next_cno = (max_no or 0) + 1

    for row_idx, row in enumerate(rows[1:], start=2):
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        results["total"] += 1
        row_data = {}
        for col_idx, field_name in reverse_map.items():
            if col_idx < len(row):
                val = row[col_idx]
                if val is not None:
                    row_data[field_name] = val
                else:
                    row_data[field_name] = ""

        # Check required fields
        missing = [f for f in required if not row_data.get(f) and row_data.get(f) != 0]
        if missing:
            cn_names = {v: k for k, v in col_map.items()}
            missing_cn = [cn_names.get(f, f) for f in missing]
            results["errors"].append({"row": row_idx, "error": f"缺少必填字段: {', '.join(missing_cn)}"})
            continue

        # Auto-assign customer_no for customers
        if module == "customers":
            if not row_data.get("customer_no") or row_data.get("customer_no") == 0:
                row_data["customer_no"] = _next_cno
                _next_cno += 1
            else:
                # User provided a custom number; ensure next auto-number stays above it
                _next_cno = max(_next_cno, int(row_data["customer_no"]) + 1)

        # Preprocess and validate
        savepoint = db.begin_nested()
        try:
            # Resolve customer name -> customer_id for orders
            if module == "orders" and "customer_name_import" in row_data:
                name = str(row_data.pop("customer_name_import")).strip()
                c = db.query(Customer).filter(
                    Customer.tenant_id == user.tenant_id,
                    Customer.name == name,
                ).first()
                if not c:
                    c = Customer(tenant_id=user.tenant_id, name=name)
                    db.add(c)
                    db.flush()
                row_data["customer_id"] = c.id
            row_data = _preprocess_row(module, row_data, db, row_idx)
            validated = SchemaClass(**row_data)
            # Customer upsert: merge by id_card if it already exists
            if module == "customers":
                id_card = validated.id_card if hasattr(validated, "id_card") else ""
                if id_card:
                    existing_c = db.query(Customer).filter(
                        Customer.tenant_id == user.tenant_id,
                        Customer.id_card == id_card,
                    ).first()
                    if existing_c:
                        vd = validated.model_dump()
                        for k, v in vd.items():
                            if k not in ("customer_no",) and v not in (None, "", 0):
                                setattr(existing_c, k, v)
                        db.flush()
                        savepoint.commit()
                        results["success"] += 1
                        continue
            obj = ModelClass(**validated.model_dump(), tenant_id=user.tenant_id)
            db.add(obj)
            db.flush()
            if module == "orders":
                auto_generate_plans(db, obj)
            savepoint.commit()
            results["success"] += 1
        except Exception as e:
            savepoint.rollback()
            err_msg = str(e)
            if "validation error" in err_msg.lower():
                err_msg = err_msg.split("\n")[1] if "\n" in err_msg else err_msg
            results["errors"].append({"row": row_idx, "error": err_msg})

    if results["success"] > 0:
        db.commit()

    return results


@router.get("/{module}/template")
def download_template(module: str):
    if module not in COLUMN_MAPPINGS:
        raise HTTPException(status_code=400, detail=f"不支持的模块: {module}")

    col_map = COLUMN_MAPPINGS[module]
    chinese_headers = list(col_map.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = MODULE_TITLES.get(module, module)

    for col_idx, header in enumerate(chinese_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) * 3, 12)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{MODULE_TITLES.get(module, module)}导入模板.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )
