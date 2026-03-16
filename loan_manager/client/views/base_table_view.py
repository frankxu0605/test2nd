"""Base table view with common CRUD UI patterns for all modules."""
from datetime import date, time

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLineEdit, QLabel, QMessageBox, QHeaderView, QDialog,
    QFormLayout, QComboBox, QDateEdit, QTimeEdit, QDoubleSpinBox, QSpinBox,
    QTextEdit, QAbstractItemView, QFileDialog, QCompleter,
)
from PyQt6.QtCore import Qt, QDate, QTime, QTimer, QStringListModel
from PyQt6.QtGui import QFont

from openpyxl import Workbook


class BaseTableView(QWidget):
    """
    Subclass and override:
        - columns: list of (header_label, dict_key, width)
        - get_form_fields() -> list of field defs
        - fetch_data(keyword) -> list[dict]
        - create_item(data) -> dict
        - update_item(item_id, data) -> dict
        - delete_item(item_id) -> dict
    """
    columns: list[tuple[str, str, int]] = []
    page_title: str = ""

    def __init__(self):
        super().__init__()
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        # Toolbar
        toolbar = QHBoxLayout()

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("搜索...")
        self.search_input.setFixedWidth(250)
        self.search_input.setFixedHeight(32)
        self.search_input.returnPressed.connect(self.load_data)
        toolbar.addWidget(self.search_input)

        search_btn = QPushButton("搜索")
        search_btn.setFixedHeight(32)
        search_btn.setStyleSheet("QPushButton { background-color: #1890ff; color: white; border: none; border-radius: 4px; padding: 0 16px; }")
        search_btn.clicked.connect(self.load_data)
        toolbar.addWidget(search_btn)

        toolbar.addStretch()

        add_btn = QPushButton("+ 新增")
        add_btn.setFixedHeight(32)
        add_btn.setStyleSheet("QPushButton { background-color: #52c41a; color: white; border: none; border-radius: 4px; padding: 0 16px; }")
        add_btn.clicked.connect(self._on_add)
        toolbar.addWidget(add_btn)

        export_btn = QPushButton("导出Excel")
        export_btn.setFixedHeight(32)
        export_btn.setStyleSheet("QPushButton { background-color: #faad14; color: white; border: none; border-radius: 4px; padding: 0 16px; }")
        export_btn.clicked.connect(self._on_export)
        toolbar.addWidget(export_btn)

        layout.addLayout(toolbar)

        # Table
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setDefaultSectionSize(36)
        self.table.setStyleSheet(
            "QTableWidget { background-color: white; border: 1px solid #e8e8e8; gridline-color: #f0f0f0; }"
            "QTableWidget::item:selected { background-color: #e6f7ff; color: black; }"
            "QHeaderView::section { background-color: #fafafa; border: none; border-bottom: 1px solid #e8e8e8; padding: 8px; font-weight: bold; }"
        )

        headers = [c[0] for c in self.columns] + ["操作"]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        for i, col in enumerate(self.columns):
            if col[2] > 0:
                self.table.setColumnWidth(i, col[2])

        layout.addWidget(self.table)

    def load_data(self):
        keyword = self.search_input.text().strip()
        try:
            data = self.fetch_data(keyword)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载数据失败: {e}")
            return

        self.table.setRowCount(len(data))
        for row_idx, item in enumerate(data):
            for col_idx, (_, key, _) in enumerate(self.columns):
                val = item.get(key, "")
                if val is None:
                    val = ""
                cell = QTableWidgetItem(str(val))
                cell.setData(Qt.ItemDataRole.UserRole, item)
                self.table.setItem(row_idx, col_idx, cell)

            # Action buttons
            action_widget = QWidget()
            action_layout = QHBoxLayout(action_widget)
            action_layout.setContentsMargins(4, 2, 4, 2)
            action_layout.setSpacing(4)

            edit_btn = QPushButton("编辑")
            edit_btn.setFixedSize(50, 26)
            edit_btn.setStyleSheet("QPushButton { background-color: #1890ff; color: white; border: none; border-radius: 3px; font-size: 12px; }")
            edit_btn.clicked.connect(lambda checked, r=row_idx: self._on_edit(r))
            action_layout.addWidget(edit_btn)

            del_btn = QPushButton("删除")
            del_btn.setFixedSize(50, 26)
            del_btn.setStyleSheet("QPushButton { background-color: #ff4d4f; color: white; border: none; border-radius: 3px; font-size: 12px; }")
            del_btn.clicked.connect(lambda checked, r=row_idx: self._on_delete(r))
            action_layout.addWidget(del_btn)

            self.table.setCellWidget(row_idx, len(self.columns), action_widget)

    def _get_row_data(self, row: int) -> dict | None:
        item = self.table.item(row, 0)
        if item:
            return item.data(Qt.ItemDataRole.UserRole)
        return None

    def _on_add(self):
        dlg = FormDialog(self, "新增", self.get_form_fields())
        if dlg.exec() == QDialog.DialogCode.Accepted:
            data = dlg.get_data()
            try:
                self.create_item(data)
                self.load_data()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"新增失败: {e}")

    def _on_edit(self, row: int):
        row_data = self._get_row_data(row)
        if not row_data:
            return
        dlg = FormDialog(self, "编辑", self.get_form_fields(), row_data)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            data = dlg.get_data()
            try:
                self.update_item(row_data["id"], data)
                self.load_data()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"更新失败: {e}")

    def _on_delete(self, row: int):
        row_data = self._get_row_data(row)
        if not row_data:
            return
        reply = QMessageBox.question(self, "确认删除", "确定要删除这条记录吗？")
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.delete_item(row_data["id"])
                self.load_data()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"删除失败: {e}")

    def _on_export(self):
        if self.table.rowCount() == 0:
            QMessageBox.information(self, "提示", "没有数据可导出")
            return

        path, _ = QFileDialog.getSaveFileName(self, "导出Excel", f"{self.page_title}.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = self.page_title or "数据"

        # Headers
        for col_idx, (header, _, _) in enumerate(self.columns):
            ws.cell(row=1, column=col_idx + 1, value=header)

        # Data
        for row_idx in range(self.table.rowCount()):
            for col_idx in range(len(self.columns)):
                item = self.table.item(row_idx, col_idx)
                ws.cell(row=row_idx + 2, column=col_idx + 1, value=item.text() if item else "")

        wb.save(path)
        QMessageBox.information(self, "成功", f"已导出到 {path}")

    # ---- Override in subclass ----
    def get_form_fields(self) -> list[dict]:
        return []

    def fetch_data(self, keyword: str) -> list[dict]:
        return []

    def create_item(self, data: dict) -> dict:
        return {}

    def update_item(self, item_id: int, data: dict) -> dict:
        return {}

    def delete_item(self, item_id: int) -> dict:
        return {}


class FormDialog(QDialog):
    """Dynamic form dialog based on field definitions."""

    def __init__(self, parent, title: str, fields: list[dict], data: dict | None = None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setMinimumWidth(420)
        self.fields = fields
        self.widgets = {}
        self._init_ui(data)

    def _init_ui(self, data: dict | None):
        layout = QVBoxLayout(self)
        form = QFormLayout()
        form.setSpacing(10)

        for field in self.fields:
            key = field["key"]
            label = field["label"]
            ftype = field.get("type", "text")
            options = field.get("options", [])

            if ftype == "text":
                w = QLineEdit()
                if data and key in data:
                    w.setText(str(data[key] or ""))
            elif ftype == "textarea":
                w = QTextEdit()
                w.setMaximumHeight(80)
                if data and key in data:
                    w.setPlainText(str(data[key] or ""))
            elif ftype == "number":
                w = QDoubleSpinBox()
                w.setMaximum(999999999.99)
                w.setDecimals(2)
                if data and key in data and data[key]:
                    w.setValue(float(data[key]))
            elif ftype == "int":
                w = QSpinBox()
                w.setMaximum(999999)
                if data and key in data and data[key]:
                    w.setValue(int(data[key]))
            elif ftype == "rate":
                w = QDoubleSpinBox()
                w.setMaximum(1.0)
                w.setDecimals(4)
                w.setSingleStep(0.0001)
                if data and key in data and data[key]:
                    w.setValue(float(data[key]))
            elif ftype == "date":
                w = QDateEdit()
                w.setCalendarPopup(True)
                w.setDisplayFormat("yyyy-MM-dd")
                if data and key in data and data[key]:
                    try:
                        parts = str(data[key]).split("-")
                        w.setDate(QDate(int(parts[0]), int(parts[1]), int(parts[2])))
                    except (ValueError, IndexError):
                        w.setDate(QDate.currentDate())
                else:
                    w.setDate(QDate.currentDate())
            elif ftype == "time":
                w = QTimeEdit()
                w.setDisplayFormat("HH:mm")
                if data and key in data and data[key]:
                    try:
                        parts = str(data[key]).split(":")
                        w.setTime(QTime(int(parts[0]), int(parts[1])))
                    except (ValueError, IndexError):
                        w.setTime(QTime.currentTime())
                else:
                    w.setTime(QTime.currentTime())
            elif ftype == "customer_search":
                w = QLineEdit()
                w.setPlaceholderText("输入客户姓名搜索...")
                w._customer_id = ""
                w._autofill = field.get("autofill", [])
                w._customer_map = {}
                display_key = field.get("display_key", "customer_name")
                if data and data.get(display_key):
                    w.setText(str(data[display_key]))
                    w._customer_id = str(data.get(key, ""))

                completer = QCompleter()
                completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
                completer.setFilterMode(Qt.MatchFlag.MatchContains)
                model = QStringListModel()
                completer.setModel(model)
                w.setCompleter(completer)
                w._completer_model = model

                timer = QTimer(self)
                timer.setSingleShot(True)
                timer.setInterval(400)

                def _make_search(line_edit, comp):
                    def _do_search():
                        kw = line_edit.text().strip()
                        if not kw:
                            return
                        try:
                            from ..api_client import api
                            customers = api.list_customers(keyword=kw)
                            line_edit._customer_map = {}
                            names = []
                            for c in customers:
                                display = f"{c['name']}  ({c.get('phone', '')})"
                                names.append(display)
                                line_edit._customer_map[display] = c
                            line_edit._completer_model.setStringList(names)
                            comp.complete()
                        except Exception:
                            pass
                    return _do_search

                timer.timeout.connect(_make_search(w, completer))
                w.textChanged.connect(lambda text, t=timer: t.start())

                def _make_activated(line_edit, form_widgets):
                    def _on_activated(text):
                        c = line_edit._customer_map.get(text)
                        if c:
                            line_edit._customer_id = str(c['id'])
                            for fill_key in line_edit._autofill:
                                if fill_key in form_widgets:
                                    target_w, target_ftype = form_widgets[fill_key]
                                    val = str(c.get(fill_key, "") or "")
                                    if target_ftype == "text":
                                        target_w.setText(val)
                                    elif target_ftype == "combo":
                                        idx = target_w.findText(val)
                                        if idx >= 0:
                                            target_w.setCurrentIndex(idx)
                    return _on_activated

                completer.activated.connect(_make_activated(w, self.widgets))
            elif ftype == "combo":
                w = QComboBox()
                w.addItems(options)
                if data and key in data:
                    idx = w.findText(str(data[key]))
                    if idx >= 0:
                        w.setCurrentIndex(idx)
            else:
                w = QLineEdit()

            self.widgets[key] = (w, ftype)
            form.addRow(label + ":", w)

        layout.addLayout(form)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedSize(80, 32)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        ok_btn = QPushButton("确定")
        ok_btn.setFixedSize(80, 32)
        ok_btn.setStyleSheet("QPushButton { background-color: #1890ff; color: white; border: none; border-radius: 4px; }")
        ok_btn.clicked.connect(self.accept)
        btn_layout.addWidget(ok_btn)

        layout.addLayout(btn_layout)

    def get_data(self) -> dict:
        result = {}
        for key, (widget, ftype) in self.widgets.items():
            if ftype == "text":
                result[key] = widget.text()
            elif ftype == "textarea":
                result[key] = widget.toPlainText()
            elif ftype in ("number", "rate"):
                result[key] = widget.value()
            elif ftype == "int":
                result[key] = widget.value()
            elif ftype == "customer_search":
                try:
                    result[key] = int(widget._customer_id) if widget._customer_id else 0
                except ValueError:
                    result[key] = 0
            elif ftype == "date":
                result[key] = widget.date().toString("yyyy-MM-dd")
            elif ftype == "time":
                result[key] = widget.time().toString("HH:mm:ss")
            elif ftype == "combo":
                result[key] = widget.currentText()
            else:
                result[key] = widget.text()
        return result
